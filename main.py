from tkinter import Tk     
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook
import json

# Choose a file

newFile = open("data.json","w")

Tk().withdraw() 
fileName = askopenfilename() 

# Open a excel file

workbook = load_workbook(filename=fileName,read_only=True)
sheet = workbook.active

# Create dictionary

objects = {}

# Get data from excel file

for row in sheet.iter_rows(min_row=1,values_only=True):

    object_name = row[0] 
    object = {
        "Size":row[1],
        "Population":row[2],
        "Countries":row[3],
        "Age":row[4],
        "Climat":row[5]
    }

    # Add data to dictionary

    objects[object_name] = object
    # print(json.dumps(objects))

# Write object to json file 

newFile.write(json.dumps(objects))
newFile.close()
